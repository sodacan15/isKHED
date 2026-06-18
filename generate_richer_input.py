from openpyxl import Workbook
from pathlib import Path


def write_sheet(ws, rows):
    for row in rows:
        ws.append(row)


def build_input_workbook(path: str):
    wb = Workbook()
    wb.remove(wb.active)

    professors = [
        [
            "prof_id", "name", "days_available", "f2f_start", "f2f_end",
            "online_start", "online_end", "days_f2f", "preference_online",
            "preference_f2f", "subjects_handled"
        ],
        ["P1", "Prof. Cruz", "Monday,Tuesday,Wednesday,Thursday,Friday", 480, 1200, 450, 1260, "Monday,Tuesday,Wednesday,Thursday,Friday", True, False, "C1,C2,C3,C4,C11"],
        ["P2", "Prof. Santos", "Monday,Tuesday,Wednesday,Thursday,Friday", 480, 1200, 450, 1260, "Monday,Tuesday,Wednesday,Thursday,Friday", False, True, "C5,C6,C7,C8,C12"],
        ["P3", "Prof. Reyes", "Monday,Tuesday,Wednesday,Thursday,Friday", 480, 1200, 450, 1260, "Monday,Tuesday,Wednesday,Thursday,Friday", True, True, "C9,C10,C13,C14"],
        ["P4", "Prof. Mendoza", "Monday,Tuesday,Wednesday,Thursday,Friday", 480, 1200, 450, 1260, "Monday,Tuesday,Wednesday,Thursday,Friday", False, True, "C15,C16,C17"],
        ["P5", "Prof. Lee", "Monday,Tuesday,Wednesday,Thursday,Friday", 480, 1200, 450, 1260, "Monday,Tuesday,Wednesday,Thursday,Friday", True, False, "C18,C19,C20"],
        ["P6", "Prof. Gomez", "Monday,Tuesday,Wednesday,Thursday,Friday", 480, 1200, 450, 1260, "Monday,Tuesday,Wednesday,Thursday,Friday", True, True, "C21,C22,C23"],
    ]

    courses = [
        [
            "course_id", "mode", "has_lab", "has_lec", "time_for_lab",
            "time_for_lec", "year_level", "units", "college_name", "hour_allocation"
        ],
        ["C1", "f2f", True, True, 90, 60, 1, 3, "CCIS", 3],
        ["C2", "online", False, True, 0, 60, 1, 3, "CCIS", 3],
        ["C3", "f2f", True, True, 90, 60, 1, 3, "CCIS", 3],
        ["C4", "online", False, True, 0, 60, 1, 3, "CCIS", 3],
        ["C5", "f2f", True, True, 90, 60, 2, 3, "CCIS", 3],
        ["C6", "online", False, True, 0, 60, 2, 3, "CCIS", 3],
        ["C7", "f2f", True, True, 90, 60, 2, 3, "CCIS", 3],
        ["C8", "online", False, True, 0, 60, 2, 3, "CCIS", 3],
        ["C9", "f2f", True, True, 90, 60, 3, 3, "CCIS", 3],
        ["C10", "online", False, True, 0, 60, 3, 3, "CCIS", 3],
        ["C11", "f2f", True, True, 90, 60, 2, 3, "CCIS", 3],
        ["C12", "online", False, True, 0, 60, 2, 3, "CCIS", 3],
        ["C13", "f2f", True, True, 90, 60, 3, 3, "CCIS", 3],
        ["C14", "online", False, True, 0, 60, 3, 3, "CCIS", 3],
        ["C15", "f2f", True, True, 90, 60, 4, 3, "CCIS", 3],
        ["C16", "online", False, True, 0, 60, 4, 3, "CCIS", 3],
        ["C17", "f2f", True, True, 90, 60, 4, 3, "CCIS", 3],
        ["C18", "online", False, True, 0, 60, 4, 3, "CCIS", 3],
        ["C19", "f2f", True, True, 90, 60, 4, 3, "CCIS", 3],
        ["C20", "online", False, True, 0, 60, 4, 3, "CCIS", 3],
        ["C21", "f2f", True, True, 90, 60, 4, 3, "CCIS", 3],
        ["C22", "online", False, True, 0, 60, 4, 3, "CCIS", 3],
        ["C23", "f2f", True, True, 90, 60, 4, 3, "CCIS", 3],
    ]

    rooms = [
        ["location_code", "location_map", "capacity", "available_days"],
        ["4th_east_wing", "4th East Wing", 40, "Monday,Tuesday,Wednesday,Thursday,Friday"],
        ["5th_south_wing", "5th South Wing", 40, "Monday,Tuesday,Wednesday,Thursday,Friday"],
        ["gymnasium", "Gymnasium", 80, "Monday,Tuesday,Wednesday,Thursday,Friday"],
        ["lab_room", "Lab Room", 30, "Monday,Tuesday,Wednesday,Thursday,Friday"],
    ]

    blocks = [
        ["year_level", "block_code", "amount_of_students", "classes"],
        [1, "1-A", 35, "C1,C2,C3,C4"],
        [1, "1-B", 35, "C1,C2,C3,C4"],
        [2, "2-A", 35, "C5,C6,C7,C8,C11,C12"],
        [3, "3-A", 35, "C9,C10,C13,C14"],
        [4, "4-A", 35, "C15,C16,C17,C18,C19,C20,C21,C22,C23"],
    ]

    students = [
        ["name", "year_level", "block", "status", "backlog", "courses"],
        ["Student 1", 1, "1-A", "regular", "", "C1,C2,C3,C4"],
        ["Student 2", 1, "1-B", "regular", "", "C1,C2,C3,C4"],
        ["Student 3", 2, "2-A", "regular", "", "C5,C6,C7,C8,C11,C12"],
        ["Student 4", 3, "3-A", "regular", "", "C9,C10,C13,C14"],
        ["Student 5", 4, "4-A", "regular", "", "C15,C16,C17,C18,C19,C20,C21,C22,C23"],
        ["Student 6", 2, "2-A", "irregular", "C6,C11", "C5,C6,C7,C8,C11,C12"],
        ["Student 7", 3, "3-A", "irregular", "C9,C13", "C9,C10,C13,C14"],
        ["Student 8", 4, "4-A", "irregular", "C18,C21", "C15,C16,C17,C18,C19,C20,C21,C22,C23"],
    ]

    power = [
        ["day"],
        ["Friday"],
        ["Thursday"],
    ]

    ws = wb.create_sheet("Professors")
    write_sheet(ws, professors)

    ws = wb.create_sheet("Courses")
    write_sheet(ws, courses)

    ws = wb.create_sheet("Rooms")
    write_sheet(ws, rooms)

    ws = wb.create_sheet("Blocks")
    write_sheet(ws, blocks)

    ws = wb.create_sheet("Students")
    write_sheet(ws, students)

    ws = wb.create_sheet("PowerOutage")
    write_sheet(ws, power)

    wb.save(path)


if __name__ == "__main__":
    out = Path("inputSheet_richer.xlsx")
    build_input_workbook(str(out))
    print(f"Created {out} with a richer multi-level dataset.")
